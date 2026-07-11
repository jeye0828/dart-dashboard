#!/usr/bin/env python3
"""여러 회사의 최근 N개년 재무제표를 DART에서 조회해 비교하고 엑셀로 저장

사용법:
    # 회사명 검색
    python3 compare_companies.py --search "삼성"

    # 3개 회사 3개년(2022~2024) 비교 -> 엑셀 저장
    python3 compare_companies.py --companies "삼성전자" "SK하이닉스" "LG전자" --year 2024 --output comparison.xlsx
"""

import argparse
import sys

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

import dart_api
from financial_analyzer import analyze_growth, analyze_period

HEADER_FILL = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
COMPANY_FILL = PatternFill(start_color="203864", end_color="203864", fill_type="solid")
BOLD = Font(bold=True)
WHITE_BOLD = Font(bold=True, color="FFFFFF")

RATIO_LABELS = [
    "매출총이익률(%)", "영업이익률(%)", "순이익률(%)", "ROA(%)", "ROE(%)",
    "유동비율(%)", "당좌비율(%)", "현금비율(%)", "부채비율(%)", "부채자산비율(%)",
    "이자보상배율(x)", "총자산회전율(x)", "영업현금흐름대비순이익(x)",
]
GROWTH_LABELS = [
    "revenue_성장률(%)", "operating_income_성장률(%)", "net_income_성장률(%)",
    "total_assets_성장률(%)", "total_equity_성장률(%)",
]


def cmd_search(args):
    api_key = dart_api.load_api_key()
    corp_list = dart_api.get_corp_code_list(api_key)
    matches = dart_api.search_company(args.search, corp_list)
    if not matches:
        print(f"'{args.search}'(으)로 검색된 회사가 없습니다.")
        return
    print(f"'{args.search}' 검색 결과 ({len(matches)}건, 상위 30건 표시):")
    for c in matches[:30]:
        listed = c["stock_code"] if c["stock_code"] else "비상장"
        print(f"  {c['corp_name']:<25} corp_code={c['corp_code']}  종목코드={listed}")


def resolve_company(api_key, corp_list, name):
    matches = dart_api.search_company(name, corp_list)
    if not matches:
        raise dart_api.DartApiError(f"'{name}' 회사를 찾을 수 없습니다.")
    chosen = matches[0]
    if len(matches) > 1:
        print(f"  ('{name}' 검색 결과 {len(matches)}건 중 '{chosen['corp_name']}'(corp_code={chosen['corp_code']}) 선택)")
    return chosen


def fetch_company_data(api_key, corp_list, name, base_year):
    company = resolve_company(api_key, corp_list, name)
    periods, fs_div = dart_api.get_company_periods(api_key, company["corp_code"], base_year)
    ratio_rows = [analyze_period(p) for p in periods]
    growth_rows = analyze_growth(periods)
    return {
        "input_name": name,
        "corp_name": company["corp_name"],
        "fs_div": fs_div,
        "periods": periods,
        "ratio_rows": ratio_rows,
        "growth_rows": growth_rows,
    }


def fmt_cell(v):
    if v is None:
        return "N/A"
    return round(v, 2)


def autofit(ws, min_width=10, max_width=32):
    widths = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            length = len(str(cell.value))
            widths[cell.column] = max(widths.get(cell.column, 0), length)
    for col, width in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = min(max(width + 2, min_width), max_width)


def build_comparison_sheet(wb, companies_data):
    ws = wb.active
    ws.title = "비교표"

    years = [p["period"] for p in companies_data[0]["periods"]]

    # 헤더 행 1: 회사명 (연도 수만큼 병합), 헤더 행 2: 연도
    ws.cell(row=1, column=1, value="지표")
    ws.cell(row=2, column=1, value="")
    col = 2
    for cd in companies_data:
        start_col = col
        for year in years:
            ws.cell(row=2, column=col, value=year)
            col += 1
        end_col = col - 1
        ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
        cell = ws.cell(row=1, column=start_col, value=cd["corp_name"])
        cell.font = WHITE_BOLD
        cell.fill = COMPANY_FILL
        cell.alignment = Alignment(horizontal="center")

    for c in range(1, col):
        ws.cell(row=2, column=c).font = BOLD
        ws.cell(row=2, column=c).fill = HEADER_FILL
        ws.cell(row=2, column=c).alignment = Alignment(horizontal="center")

    row = 3
    for label in RATIO_LABELS:
        ws.cell(row=row, column=1, value=label).font = BOLD
        col = 2
        for cd in companies_data:
            for ratio_row in cd["ratio_rows"]:
                ws.cell(row=row, column=col, value=fmt_cell(ratio_row.get(label)))
                col += 1
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="[ 성장성 (YoY, %) ]").font = BOLD
    row += 1

    growth_periods = [g["period"] for g in companies_data[0]["growth_rows"]]
    ws.cell(row=row, column=1, value="")
    col = 2
    for cd in companies_data:
        for gp in growth_periods:
            ws.cell(row=row, column=col, value=gp)
            col += 1
    row += 1

    for label in GROWTH_LABELS:
        ws.cell(row=row, column=1, value=label).font = BOLD
        col = 2
        for cd in companies_data:
            for growth_row in cd["growth_rows"]:
                ws.cell(row=row, column=col, value=fmt_cell(growth_row.get(label)))
                col += 1
        row += 1

    ws.freeze_panes = "B3"
    autofit(ws)


def build_company_sheet(wb, cd):
    title = cd["corp_name"][:28] or cd["input_name"][:28]
    ws = wb.create_sheet(title=title)
    ws.cell(row=1, column=1, value=f"{cd['corp_name']} (재무제표 기준: {'연결' if cd['fs_div']=='CFS' else '개별'})").font = BOLD

    row = 3
    ws.cell(row=row, column=1, value="계정")
    for i, p in enumerate(cd["periods"], start=2):
        ws.cell(row=row, column=i, value=p["period"]).font = BOLD
        ws.cell(row=row, column=i).fill = HEADER_FILL
    ws.cell(row=row, column=1).font = BOLD
    ws.cell(row=row, column=1).fill = HEADER_FILL
    row += 1

    raw_fields = [
        ("매출액", "revenue"), ("매출원가", "cogs"), ("영업이익", "operating_income"),
        ("당기순이익", "net_income"), ("이자비용/금융비용", "interest_expense"),
        ("자산총계", "total_assets"), ("유동자산", "current_assets"), ("재고자산", "inventory"),
        ("현금및현금성자산", "cash"), ("부채총계", "total_liabilities"), ("유동부채", "current_liabilities"),
        ("자본총계", "total_equity"), ("영업활동현금흐름", "operating_cash_flow"),
    ]
    for label, key in raw_fields:
        ws.cell(row=row, column=1, value=label)
        for i, p in enumerate(cd["periods"], start=2):
            v = p.get(key)
            ws.cell(row=row, column=i, value=round(v, 0) if v is not None else "N/A")
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="[ 재무비율 ]").font = BOLD
    row += 1
    for label in RATIO_LABELS:
        ws.cell(row=row, column=1, value=label)
        for i, rr in enumerate(cd["ratio_rows"], start=2):
            ws.cell(row=row, column=i, value=fmt_cell(rr.get(label)))
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="[ 성장성 (YoY, %) ]").font = BOLD
    row += 1
    for label in GROWTH_LABELS:
        ws.cell(row=row, column=1, value=label)
        for i, gr in enumerate(cd["growth_rows"], start=2):
            ws.cell(row=row, column=i, value=fmt_cell(gr.get(label)))
        row += 1

    ws.freeze_panes = "B4"
    autofit(ws)


def cmd_compare(args):
    if len(args.companies) < 2:
        print("오류: 비교할 회사를 2개 이상 입력해주세요 (--companies A B C)", file=sys.stderr)
        sys.exit(1)

    api_key = dart_api.load_api_key()
    print("DART 기업 코드 목록 로딩 중...")
    corp_list = dart_api.get_corp_code_list(api_key)

    companies_data = []
    for name in args.companies:
        print(f"조회 중: {name} ({args.year - 2}~{args.year})")
        try:
            cd = fetch_company_data(api_key, corp_list, name, args.year)
        except dart_api.DartApiError as e:
            print(f"  오류: {e}", file=sys.stderr)
            sys.exit(1)
        companies_data.append(cd)

    wb = Workbook()
    build_comparison_sheet(wb, companies_data)
    for cd in companies_data:
        build_company_sheet(wb, cd)

    wb.save(args.output)
    print(f"\n엑셀 저장 완료: {args.output}")


def main():
    parser = argparse.ArgumentParser(description="DART 기반 회사 재무제표 비교 분석기")
    parser.add_argument("--search", help="회사명으로 DART 등록 회사 검색")
    parser.add_argument("--companies", nargs="+", help="비교할 회사명 목록 (2개 이상)")
    parser.add_argument("--year", type=int, help="기준 연도 (최근 연도, 이전 2개년도 함께 조회됨)")
    parser.add_argument("--output", default="comparison.xlsx", help="저장할 엑셀 파일 경로")
    args = parser.parse_args()

    try:
        if args.search:
            cmd_search(args)
        elif args.companies:
            if not args.year:
                print("오류: --companies 사용 시 --year 를 지정해주세요.", file=sys.stderr)
                sys.exit(1)
            cmd_compare(args)
        else:
            parser.print_help()
    except dart_api.DartApiError as e:
        print(f"오류: {e}", file=sys.stderr)
        sys.exit(1)


if __name__ == "__main__":
    main()
